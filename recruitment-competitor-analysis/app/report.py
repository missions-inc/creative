"""Excelレポートの生成

シート構成:
  ① サマリー         … クライアント概要・給与相場比較・分析条件
  ② 競合求人比較     … クライアント vs 競合求人（除外分は理由付きで下部に掲載）
  ③ 給与相場データ   … 収集サンプル全件と、数式による相場統計（下限/中央値/上限）

相場統計・時給換算はすべてExcel数式で算出するため、
データを差し替えれば再計算される。
"""

from datetime import date
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .config import ALLOWED_DOMAINS, FRESHNESS_DAYS, STANDARD_MONTHLY_HOURS
from .filters import ExcludedJob, MinWageInfo
from .schemas import ClientJob, CompetitorJob, SalarySample

# ---------------------------------------------------------------- スタイル定義
FONT_NAME = "Yu Gothic"  # 日本語ビジネス文書の標準フォント

NAVY = "1F4E79"
LIGHT_BLUE = "DDEBF7"
LIGHT_GRAY = "F2F2F2"
CLIENT_FILL = "FFF2CC"   # クライアント行のハイライト
EXCLUDE_FILL = "FCE4E4"  # 除外求人

F_TITLE = Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF")
F_SECTION = Font(name=FONT_NAME, size=11, bold=True, color=NAVY)
F_HEADER = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
F_BODY = Font(name=FONT_NAME, size=10)
F_BODY_BOLD = Font(name=FONT_NAME, size=10, bold=True)
F_SMALL = Font(name=FONT_NAME, size=9, color="595959")
F_INPUT = Font(name=FONT_NAME, size=10, color="0000FF")   # ハードコード入力値（青）
F_LINK = Font(name=FONT_NAME, size=10, color="008000")    # 他シート参照（緑）

FILL_TITLE = PatternFill("solid", fgColor=NAVY)
FILL_HEADER = PatternFill("solid", fgColor=NAVY)
FILL_SECTION = PatternFill("solid", fgColor=LIGHT_BLUE)
FILL_STRIPE = PatternFill("solid", fgColor=LIGHT_GRAY)
FILL_CLIENT = PatternFill("solid", fgColor=CLIENT_FILL)
FILL_EXCLUDE = PatternFill("solid", fgColor=EXCLUDE_FILL)

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
YEN = '#,##0"円"'

SHEET1 = "サマリー"
SHEET2 = "競合求人比較"
SHEET3 = "給与相場データ"

UNIT_ORDER = ["時給", "日給", "月給", "年収"]


def _set(ws: Worksheet, cell: str, value, font=F_BODY, fill=None, align=None,
         number_format=None, border=False):
    c = ws[cell]
    c.value = value
    c.font = font
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if number_format:
        c.number_format = number_format
    if border:
        c.border = BORDER
    return c


def _section(ws: Worksheet, row: int, text: str, span: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    _set(ws, f"A{row}", text, font=F_SECTION, fill=FILL_SECTION)


def _salary_text(min_v: Optional[int], max_v: Optional[int], unit: str) -> str:
    if not min_v:
        return "記載なし"
    if max_v and max_v != min_v:
        return f"{unit} {min_v:,}円 〜 {max_v:,}円"
    return f"{unit} {min_v:,}円"


# ---------------------------------------------------------------- シート③
def _build_market_sheet(
    ws: Worksheet, samples: Sequence[SalarySample], dropped_count: int
) -> Dict[str, Dict[str, str]]:
    """給与相場データシートを作成し、単位ごとの統計セル番地を返す。

    返り値: {単位: {"count": "B5", "low": "C5", "mid": "D5", "high": "E5"}}
    """
    ws.merge_cells("A1:I1")
    _set(ws, "A1", "③ 給与相場データ（同職種・近隣エリアの求人サンプル）",
         font=F_TITLE, fill=FILL_TITLE, align=CENTER)
    ws.row_dimensions[1].height = 24

    # --- 統計サマリー ---
    _set(ws, "A3", "■ 相場統計（下記サンプルから数式で算出）", font=F_SECTION)
    stat_header_row = 4
    headers = ["給与単位", "サンプル件数", "相場下限", "相場中央値", "相場上限"]
    for i, h in enumerate(headers, start=1):
        _set(ws, f"{get_column_letter(i)}{stat_header_row}", h,
             font=F_HEADER, fill=FILL_HEADER, align=CENTER, border=True)

    # データはこの下に単位ごとに連続配置する。まず配置行を計算する。
    ordered: List[SalarySample] = []
    unit_ranges: Dict[str, Tuple[int, int]] = {}
    units_present = [u for u in UNIT_ORDER if any(s.salary_unit == u for s in samples)]

    stats_rows = max(len(units_present), 1)
    note_row = stat_header_row + stats_rows + 1
    data_title_row = note_row + 2
    data_header_row = data_title_row + 1
    data_start = data_header_row + 1

    cursor = data_start
    for unit in units_present:
        group = [s for s in samples if s.salary_unit == unit]
        group.sort(key=lambda s: s.salary_min)
        unit_ranges[unit] = (cursor, cursor + len(group) - 1)
        ordered.extend(group)
        cursor += len(group)

    # --- 統計行（数式） ---
    stats_cells: Dict[str, Dict[str, str]] = {}
    if units_present:
        for i, unit in enumerate(units_present):
            r = stat_header_row + 1 + i
            r1, r2 = unit_ranges[unit]
            _set(ws, f"A{r}", unit, font=F_BODY_BOLD, align=CENTER, border=True)
            _set(ws, f"B{r}", f"=COUNT(E{r1}:E{r2})", font=F_BODY, align=CENTER,
                 number_format='0"件"', border=True)
            _set(ws, f"C{r}", f"=MIN(E{r1}:E{r2})", font=F_BODY_BOLD,
                 number_format=YEN, border=True)
            _set(ws, f"D{r}", f"=MEDIAN(G{r1}:G{r2})", font=F_BODY_BOLD,
                 number_format=YEN, border=True)
            _set(ws, f"E{r}", f"=MAX(E{r1}:E{r2},F{r1}:F{r2})", font=F_BODY_BOLD,
                 number_format=YEN, border=True)
            stats_cells[unit] = {
                "count": f"B{r}", "low": f"C{r}", "mid": f"D{r}", "high": f"E{r}",
            }
    else:
        _set(ws, f"A{stat_header_row + 1}",
             "サンプルを取得できませんでした。検索条件を変えて再実行してください。",
             font=F_BODY)

    _set(ws, f"A{note_row}",
         "※ 相場下限＝各求人の下限給与の最小値 ／ 相場中央値＝各求人の代表値"
         "（下限と上限の中間）の中央値 ／ 相場上限＝各求人の上限給与の最大値",
         font=F_SMALL)

    # --- データ一覧 ---
    _set(ws, f"A{data_title_row}",
         f"■ 収集サンプル一覧（{len(ordered)}件"
         + (f"、ほか{dropped_count}件を最低賃金未満・単位不明のため除外" if dropped_count else "")
         + "）",
         font=F_SECTION)

    data_headers = ["No", "求人（企業名・タイトル）", "勤務地", "媒体",
                    "給与下限", "給与上限", "代表値", "単位", "出典URL"]
    for i, h in enumerate(data_headers, start=1):
        _set(ws, f"{get_column_letter(i)}{data_header_row}", h,
             font=F_HEADER, fill=FILL_HEADER, align=CENTER, border=True)

    for idx, s in enumerate(ordered):
        r = data_start + idx
        stripe = FILL_STRIPE if idx % 2 else None
        _set(ws, f"A{r}", idx + 1, font=F_BODY, fill=stripe, align=CENTER, border=True)
        _set(ws, f"B{r}", s.label, font=F_BODY, fill=stripe, align=WRAP, border=True)
        _set(ws, f"C{r}", s.location, font=F_BODY, fill=stripe, align=WRAP, border=True)
        _set(ws, f"D{r}", s.source_media, font=F_BODY, fill=stripe, align=CENTER, border=True)
        _set(ws, f"E{r}", s.salary_min, font=F_BODY, fill=stripe,
             number_format=YEN, border=True)
        _set(ws, f"F{r}", s.salary_max, font=F_BODY, fill=stripe,
             number_format=YEN, border=True)
        _set(ws, f"G{r}", f'=IF(F{r}="",E{r},(E{r}+F{r})/2)', font=F_BODY,
             fill=stripe, number_format=YEN, border=True)
        _set(ws, f"H{r}", s.salary_unit, font=F_BODY, fill=stripe, align=CENTER, border=True)
        _set(ws, f"I{r}", s.url, font=F_SMALL, fill=stripe, align=WRAP, border=True)

    widths = [5, 34, 18, 12, 13, 13, 13, 8, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{data_start}"
    return stats_cells


# ---------------------------------------------------------------- シート②
def _hourly_formula(r: int, monthly_hours_ref: str) -> str:
    """下限給与の時給換算式（I=下限, K=単位）"""
    return (
        f'=IF(I{r}="","-",'
        f'IF(K{r}="時給",I{r},'
        f'IF(K{r}="日給",I{r}/8,'
        f'IF(K{r}="月給",I{r}/{monthly_hours_ref},'
        f'IF(K{r}="年収",I{r}/12/{monthly_hours_ref},"-")))))'
    )


def _competitor_row(ws: Worksheet, r: int, no: str, kind: str, media: str,
                    company: str, title: str, emp: str, loc: str,
                    salary_text: str, s_min, s_max, unit: str,
                    pr: str, benefits: str, updated: str, url: str,
                    monthly_hours_ref: str, fill=None, bold=False):
    font = F_BODY_BOLD if bold else F_BODY
    values = [no, kind, media, company, title, emp, loc, salary_text,
              s_min, s_max, unit, None, pr, benefits, updated, url]
    for i, v in enumerate(values, start=1):
        col = get_column_letter(i)
        cell = _set(ws, f"{col}{r}", v, font=font, fill=fill, border=True)
        if i in (1, 2, 3, 6, 11, 15):
            cell.alignment = CENTER
        elif i in (5, 7, 8, 13, 14, 16):
            cell.alignment = WRAP
        if i in (9, 10):
            cell.number_format = YEN
    # 時給換算（数式）
    _set(ws, f"L{r}", _hourly_formula(r, monthly_hours_ref), font=font, fill=fill,
         number_format='#,##0"円"', border=True, align=CENTER)


def _build_competitor_sheet(
    ws: Worksheet,
    client_job: ClientJob,
    competitors: Sequence[CompetitorJob],
    excluded: Sequence[ExcludedJob],
    monthly_hours_ref: str,
):
    ncols = 16
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    _set(ws, "A1", "② 競合求人比較（クライアント vs 近隣競合）",
         font=F_TITLE, fill=FILL_TITLE, align=CENTER)
    ws.row_dimensions[1].height = 24

    headers = ["No", "区分", "媒体", "企業・事業所名", "職種・タイトル", "雇用形態",
               "勤務地", "給与（表記）", "給与下限", "給与上限", "単位",
               "時給換算(下限)", "PRポイント", "福利厚生・制度", "更新日", "URL"]
    for i, h in enumerate(headers, start=1):
        _set(ws, f"{get_column_letter(i)}3", h, font=F_HEADER, fill=FILL_HEADER,
             align=CENTER, border=True)

    # クライアント行
    r = 4
    _competitor_row(
        ws, r, "-", "クライアント", "-", client_job.company or "（クライアント）",
        client_job.job_title, client_job.employment_type,
        f"{client_job.prefecture}{client_job.city}",
        _salary_text(client_job.salary_min, client_job.salary_max, client_job.salary_unit),
        client_job.salary_min, client_job.salary_max, client_job.salary_unit,
        "・" + "\n・".join(client_job.pr_points) if client_job.pr_points else "",
        "・" + "\n・".join(client_job.benefits) if client_job.benefits else "",
        "-", "", monthly_hours_ref, fill=FILL_CLIENT, bold=True,
    )

    for idx, job in enumerate(competitors):
        r += 1
        stripe = FILL_STRIPE if idx % 2 else None
        _competitor_row(
            ws, r, str(idx + 1), "競合", job.source_media, job.company, job.title,
            job.employment_type, job.location, job.salary_text,
            job.salary_min, job.salary_max, job.salary_unit,
            "・" + "\n・".join(job.pr_points) if job.pr_points else "",
            "・" + "\n・".join(job.benefits) if job.benefits else "",
            job.updated_date or "不明", job.url, monthly_hours_ref, fill=stripe,
        )

    r += 2
    if excluded:
        _set(ws, f"A{r}",
             f"■ 除外した求人（{len(excluded)}件） … 除外ルール: "
             f"最低賃金未満 ／ 更新日が{FRESHNESS_DAYS}日以上前",
             font=F_SECTION)
        r += 1
        ex_headers = ["No", "媒体", "企業・事業所名", "職種・タイトル", "給与（表記）", "除外理由"]
        for i, h in enumerate(ex_headers, start=1):
            _set(ws, f"{get_column_letter(i)}{r}", h, font=F_HEADER,
                 fill=FILL_HEADER, align=CENTER, border=True)
        for i, ex in enumerate(excluded):
            r += 1
            for col, v in zip("ABCDEF", [str(i + 1), ex.job.source_media, ex.job.company,
                                         ex.job.title, ex.job.salary_text, ex.reason]):
                _set(ws, f"{col}{r}", v, font=F_BODY, fill=FILL_EXCLUDE,
                     align=WRAP, border=True)

    widths = [5, 11, 12, 22, 26, 11, 18, 20, 12, 12, 7, 12, 32, 32, 11, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------- シート①
def _build_summary_sheet(
    ws: Worksheet,
    client_job: ClientJob,
    stats_cells: Dict[str, Dict[str, str]],
    min_wage: MinWageInfo,
    today: date,
    sample_count: int,
    dropped_count: int,
    competitor_count: int,
    excluded_count: int,
) -> str:
    """サマリーシートを作成し、月平均所定労働時間セルの参照文字列を返す"""
    span = 8
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    _set(ws, "A1", "採用競合・給与相場 分析レポート", font=F_TITLE,
         fill=FILL_TITLE, align=CENTER)
    ws.row_dimensions[1].height = 28
    _set(ws, "A2", f"作成日: {today.strftime('%Y年%m月%d日')}", font=F_SMALL)

    # --- クライアント求人概要 ---
    _section(ws, 4, "■ クライアント求人概要", span)
    rows = [
        ("施設・企業名", client_job.company or "（記載なし）"),
        ("募集職種", f"{client_job.job_title}（{client_job.job_category}）"),
        ("雇用形態", client_job.employment_type),
        ("勤務地", f"{client_job.prefecture}{client_job.city}"),
        ("提示給与", _salary_text(client_job.salary_min, client_job.salary_max,
                                  client_job.salary_unit)),
        ("勤務時間", client_job.work_hours or "（記載なし）"),
        ("PRポイント", " ／ ".join(client_job.pr_points) or "（記載なし）"),
        ("福利厚生・制度", " ／ ".join(client_job.benefits) or "（記載なし）"),
    ]
    r = 5
    for label, value in rows:
        _set(ws, f"A{r}", label, font=F_BODY_BOLD, fill=FILL_STRIPE, border=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=span)
        _set(ws, f"B{r}", value, font=F_BODY, align=WRAP, border=True)
        for col in range(3, span + 1):
            ws.cell(row=r, column=col).border = BORDER
        r += 1

    # --- 給与相場比較 ---
    r += 1
    _section(ws, r, "■ 給与相場比較（③給与相場データシートより自動算出）", span)
    r += 1
    header_row = r
    headers = ["給与単位", "件数", "相場下限", "相場中央値", "相場上限",
               "クライアント下限", "クライアント上限", "相場ポジション"]
    for i, h in enumerate(headers, start=1):
        _set(ws, f"{get_column_letter(i)}{header_row}", h, font=F_HEADER,
             fill=FILL_HEADER, align=CENTER, border=True)

    if stats_cells:
        for unit, cells in stats_cells.items():
            r += 1
            _set(ws, f"A{r}", unit, font=F_BODY_BOLD, align=CENTER, border=True)
            _set(ws, f"B{r}", f"='{SHEET3}'!{cells['count']}", font=F_LINK,
                 align=CENTER, number_format='0"件"', border=True)
            _set(ws, f"C{r}", f"='{SHEET3}'!{cells['low']}", font=F_LINK,
                 number_format=YEN, border=True)
            _set(ws, f"D{r}", f"='{SHEET3}'!{cells['mid']}", font=F_LINK,
                 number_format=YEN, border=True)
            _set(ws, f"E{r}", f"='{SHEET3}'!{cells['high']}", font=F_LINK,
                 number_format=YEN, border=True)
            if unit == client_job.salary_unit and client_job.salary_min:
                _set(ws, f"F{r}", client_job.salary_min, font=F_INPUT,
                     number_format=YEN, border=True)
                _set(ws, f"G{r}", client_job.salary_max or client_job.salary_min,
                     font=F_INPUT, number_format=YEN, border=True)
                # 代表値（下限と上限の中間）を相場と比較して判定
                mid = f"(F{r}+G{r})/2"
                _set(ws, f"H{r}",
                     f'=IF({mid}>=E{r},"◎ 相場上限級",'
                     f'IF({mid}>=D{r},"○ 中央値以上",'
                     f'IF({mid}>=C{r},"△ 中央値未満","× 相場下限未満")))',
                     font=F_BODY_BOLD, align=CENTER, border=True)
            else:
                for col in ("F", "G", "H"):
                    _set(ws, f"{col}{r}", "-", font=F_BODY, align=CENTER, border=True)
    else:
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
        _set(ws, f"A{r}", "相場サンプルを取得できなかったため、比較表は作成されていません。",
             font=F_BODY, border=True)

    r += 1
    _set(ws, f"A{r}",
         "※ 相場ポジションはクライアント提示給与の代表値（下限と上限の中間）と相場統計の比較です。",
         font=F_SMALL)

    # --- 分析条件・出典 ---
    r += 2
    _section(ws, r, "■ 分析条件・出典", span)
    monthly_hours_row = None
    conditions = [
        ("データ取得日", today.strftime("%Y年%m月%d日")),
        ("対象媒体", "Indeed ／ ジョブメドレー ／ ハローワーク"
                     f"（検索対象ドメイン: {', '.join(ALLOWED_DOMAINS)}）"),
        ("除外ルール①", f"最低賃金（{min_wage.hourly:,}円/時）を下回る求人を除外"
                        f"　※出典: {min_wage.source}"),
        ("除外ルール②", f"更新日が{FRESHNESS_DAYS}日以上前の求人を除外"
                        "（更新日不明の求人は注記の上で保持）"),
        ("相場サンプル", f"採用 {sample_count}件（除外 {dropped_count}件）"),
        ("競合求人", f"採用 {competitor_count}件（除外 {excluded_count}件）"),
        ("月平均所定労働時間", STANDARD_MONTHLY_HOURS),
    ]
    for label, value in conditions:
        r += 1
        _set(ws, f"A{r}", label, font=F_BODY_BOLD, fill=FILL_STRIPE, border=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=span)
        if label == "月平均所定労働時間":
            monthly_hours_row = r
            _set(ws, f"B{r}", value, font=F_INPUT, border=True,
                 number_format='0.0"時間/月"')
        else:
            _set(ws, f"B{r}", value, font=F_BODY, align=WRAP, border=True)
        for col in range(3, span + 1):
            ws.cell(row=r, column=col).border = BORDER
    r += 1
    _set(ws, f"A{r}",
         "※ 月平均所定労働時間は月給→時給換算の前提値（週40時間×52週÷12ヶ月）。"
         "青字セルを変更すると②シートの時給換算が再計算されます。日給は8時間/日で換算。",
         font=F_SMALL)

    ws.column_dimensions["A"].width = 20
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 15

    return f"'{SHEET1}'!$B${monthly_hours_row}"


# ---------------------------------------------------------------- エントリポイント
def build_report(
    output_path: str,
    client_job: ClientJob,
    competitors: Sequence[CompetitorJob],
    excluded: Sequence[ExcludedJob],
    samples: Sequence[SalarySample],
    dropped_sample_count: int,
    min_wage: MinWageInfo,
    today: date,
):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = SHEET1
    ws2 = wb.create_sheet(SHEET2)
    ws3 = wb.create_sheet(SHEET3)

    stats_cells = _build_market_sheet(ws3, samples, dropped_sample_count)
    monthly_hours_ref = _build_summary_sheet(
        ws1, client_job, stats_cells, min_wage, today,
        sample_count=len(samples), dropped_count=dropped_sample_count,
        competitor_count=len(competitors), excluded_count=len(excluded),
    )
    _build_competitor_sheet(ws2, client_job, competitors, excluded, monthly_hours_ref)

    wb.save(output_path)
